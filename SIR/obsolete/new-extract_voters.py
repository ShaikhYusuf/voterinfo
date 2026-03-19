"""
Maharashtra Voter List PDF Extractor (OCR-based)
=================================================
This script uses OCR to extract voter data from scanned Maharashtra
Election Commission voter list PDFs.

HOW IT WORKS:
- The PDF is image-based (scanned), so pdfplumber cannot extract text directly.
- We convert each page to an image and run Tesseract OCR on it.
- Voter IDs (e.g. XIC5595830, JGH0652115) are reliably extracted by OCR.
- Age and gender are extracted from the "वय : XX लिंग : XX" line pattern.
- Marathi names require the 'mar' Tesseract language pack for accuracy.
  Without it, names appear garbled — install instructions below.

REQUIREMENTS:
    pip install pdf2image pytesseract openpyxl

    # Tesseract OCR engine (required):
    # Ubuntu/Debian:
    #   sudo apt install tesseract-ocr tesseract-ocr-mar poppler-utils
    # macOS:
    #   brew install tesseract tesseract-lang
    # Windows:
    #   Download from: https://github.com/UB-Mannheim/tesseract/wiki
    #   Download mar.traineddata from: https://github.com/tesseract-ocr/tessdata
    #   Place in: C:/Program Files/Tesseract-OCR/tessdata/

USAGE:
    # With Marathi language (recommended - extracts proper names):
    python extract_voters.py voter_list.pdf --lang mar+eng

    # Without Marathi pack (Voter ID, Age, Gender only):
    python extract_voters.py voter_list.pdf --lang eng

    # Custom output file:
    python extract_voters.py voter_list.pdf --lang mar+eng --out output.xlsx

    # Specific page range only:
    python extract_voters.py voter_list.pdf --lang mar+eng --pages 3-20
"""

import re
import sys
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from pdf2image import convert_from_path
    import pytesseract
except ImportError:
    print("ERROR: Missing dependencies. Run: pip install pdf2image pytesseract openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────
# GENDER DETECTION
# ─────────────────────────────────────────────
# When using eng-only OCR, Marathi words get transliterated:
# पुरुष (Male)   → gee, Gor, gem, Goa, yew, gow, Gee, ger
# महिला (Female) → Hfg, afg, aft, afr, niger, after, HfgeT

def detect_gender_eng_ocr(ocr_word):
    if re.search(r'(?:Hfg|afg|nig|aft|Hfr|afr)', ocr_word):
        return 'महिला (Female)'
    if re.search(r'(?:gee|Gor|gem|Goa|yew|gow|Gee|ger)', ocr_word):
        return 'पुरुष (Male)'
    return ''

def detect_gender_mar_ocr(ocr_word):
    if 'महिला' in ocr_word:
        return 'महिला (Female)'
    if 'पुरुष' in ocr_word:
        return 'पुरुष (Male)'
    return ''


# ─────────────────────────────────────────────
# PAGE PARSER
# ─────────────────────────────────────────────

def parse_page(text, page_num, lang):
    """
    Parse voter entries from OCR text of a single page.
    Voter IDs are used as delimiters to split text into per-voter blocks.
    """
    voters = []
    voter_id_re = re.compile(r'\b([A-Z]{2,3}[0-9]{6,10})\b')
    matches = list(voter_id_re.finditer(text))
    use_marathi = 'mar' in lang

    for i, match in enumerate(matches):
        voter_id  = match.group(1)
        block_end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        block     = text[match.start(): block_end]

        voter = {
            'serial':        '',
            'voter_id':      voter_id,
            'name':          '',
            'relative_name': '',
            'house_no':      '',
            'age':           '',
            'gender':        '',
            'page':          page_num,
        }

        # Age + Gender: "वय : 52 लिंग : पुरुष"  (or transliterated eng version)
        ag_re = re.compile(
            r'(?:वय|aa|aq|wa)\s*[:\|]\s*(\d+)\s+'
            r'(?:लिंग|fer|fein|fait|fet|fart|fe|fem)\s*[:\|]\s*(\S+)',
            re.IGNORECASE
        )
        ag = ag_re.search(block)
        if ag:
            voter['age']    = ag.group(1)
            voter['gender'] = (detect_gender_mar_ocr(ag.group(2)) if use_marathi
                               else detect_gender_eng_ocr(ag.group(2)))

        # Name fields — only meaningful with Marathi OCR pack
        if use_marathi:
            nm = re.search(r'नाव\s*:\s*(.+)', block)
            if nm:
                voter['name'] = nm.group(1).strip()

            rel = re.search(r'(?:वडिलांचे नाव|पतीचे नाव|आईचे नाव|इतर)\s*:\s*(.+)', block)
            if rel:
                voter['relative_name'] = rel.group(1).strip()

            hn = re.search(r'घर क्रमांक\s*:\s*(.+)', block)
            if hn:
                voter['house_no'] = hn.group(1).strip()

        # Serial number (appears just before the Voter ID)
        pre = text[max(0, match.start() - 60): match.start()]
        sr  = re.search(r'(\d+)\s*$', pre.strip())
        if sr:
            val = int(sr.group(1))
            if val < 10000:  # sanity check
                voter['serial'] = val

        voters.append(voter)

    return voters


# ─────────────────────────────────────────────
# EXCEL OUTPUT
# ─────────────────────────────────────────────

def save_to_excel(voters, output_path, pdf_path, lang):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Voter List"

    thin     = Side(style='thin', color="AAAAAA")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    alt_fill = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
    h_font   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    d_font   = Font(name="Arial", size=10)

    # Title rows
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Maharashtra Voter List — Source: {pdf_path}"
    ws["A1"].font      = Font(bold=True, name="Arial", size=12, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    note = ("Full data with names" if 'mar' in lang
            else "Voter ID/Age/Gender only — install Marathi OCR pack for names (--lang mar+eng)")
    ws["A2"] = f"Total voters: {len(voters)} | OCR: {lang} | {note}"
    ws["A2"].font      = Font(italic=True, name="Arial", size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Sr. No.", "Voter ID", "नाव (Name)", "वडिलांचे/पतीचे नाव",
               "घर क्रमांक", "लिंग (Gender)", "वय (Age)", "PDF Page"]
    widths  = [10, 18, 30, 30, 20, 20, 10, 10]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    ws.row_dimensions[3].height = 28
    ws.freeze_panes = "A4"

    for i, v in enumerate(voters):
        row  = i + 4
        fill = alt_fill if i % 2 == 0 else PatternFill()
        vals = [v['serial'], v['voter_id'], v['name'], v['relative_name'],
                v['house_no'], v['gender'], v['age'], v['page']]
        for c, val in enumerate(vals, 1):
            cell           = ws.cell(row=row, column=c, value=val)
            cell.font      = d_font
            cell.border    = border
            cell.fill      = fill
            cell.alignment = Alignment(
                horizontal="center" if c in [1, 6, 7, 8] else "left",
                vertical="center"
            )

    # Footer
    fr = len(voters) + 4
    ws.merge_cells(f"A{fr}:F{fr}")
    ws[f"A{fr}"]           = "एकूण मतदार (Total Voters)"
    ws[f"A{fr}"].font      = Font(bold=True, color="FFFFFF", name="Arial")
    ws[f"A{fr}"].fill      = h_fill
    ws[f"A{fr}"].alignment = Alignment(horizontal="center")
    ws[f"G{fr}"]           = f"=COUNTA(G4:G{fr - 1})"
    ws[f"G{fr}"].font      = Font(bold=True, color="FFFFFF", name="Arial")
    ws[f"G{fr}"].fill      = h_fill

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"]      = "Gender Summary"
    ws2["A1"].font = Font(bold=True, name="Arial", size=12)
    male   = sum(1 for v in voters if 'Male'   in v['gender'])
    female = sum(1 for v in voters if 'Female' in v['gender'])
    other  = len(voters) - male - female
    for r, (lbl, cnt) in enumerate([
        ("पुरुष (Male)",   male),
        ("महिला (Female)", female),
        ("Unknown",        other),
        ("Total",          len(voters))
    ], start=2):
        ws2.cell(row=r, column=1, value=lbl).font = d_font
        ws2.cell(row=r, column=2, value=cnt).font = Font(bold=True, name="Arial")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 10

    wb.save(output_path)
    return male, female, other


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Extract Maharashtra voter list from scanned PDF to Excel"
    )
    parser.add_argument("pdf",           help="Path to voter list PDF file")
    parser.add_argument("--lang",        default="eng",
                        help="Tesseract OCR language. Use 'mar+eng' if Marathi pack installed (default: eng)")
    parser.add_argument("--out",         default="voters_output.xlsx",
                        help="Output Excel filename (default: voters_output.xlsx)")
    parser.add_argument("--dpi",         default=150, type=int,
                        help="Image resolution for OCR. Higher=better but slower (default: 150)")
    parser.add_argument("--pages",       default=None,
                        help="Page range to process, e.g. '3-20' (default: all pages)")
    args = parser.parse_args()

    first_page = last_page = None
    if args.pages:
        parts = args.pages.split('-')
        first_page = int(parts[0])
        last_page  = int(parts[1]) if len(parts) > 1 else int(parts[0])

    print(f"\n📄 PDF     : {args.pdf}")
    print(f"🔤 Language: {args.lang}")
    print(f"🖼  DPI     : {args.dpi}")
    if first_page:
        print(f"📑 Pages   : {first_page}–{last_page}")
    print("\nConverting pages to images...")

    kwargs = {'dpi': args.dpi}
    if first_page: kwargs['first_page'] = first_page
    if last_page:  kwargs['last_page']  = last_page

    images = convert_from_path(args.pdf, **kwargs)
    print(f"   {len(images)} page(s) loaded.\n")

    all_voters   = []
    start_offset = (first_page - 1) if first_page else 0

    for idx, img in enumerate(images):
        page_num = start_offset + idx + 1
        print(f"  Page {page_num:3d} ... ", end="", flush=True)
        text   = pytesseract.image_to_string(img, lang=args.lang)
        voters = parse_page(text, page_num, args.lang)
        print(f"{len(voters)} voters")
        all_voters.extend(voters)

    print(f"\n✅ Total voters extracted : {len(all_voters)}")
    print("💾 Saving to Excel...")
    male, female, other = save_to_excel(all_voters, args.out, args.pdf, args.lang)
    print(f"   पुरुष (Male)   : {male}")
    print(f"   महिला (Female) : {female}")
    print(f"   Unknown gender : {other}")
    print(f"\n✅ Saved: {args.out}\n")


if __name__ == "__main__":
    main()
